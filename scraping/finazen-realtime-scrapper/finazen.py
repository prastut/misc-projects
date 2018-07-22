#!/usr/bin/env python
import re
import urlparse
import pickle
import os
from selenium import webdriver
from bs4 import BeautifulSoup
import time 
import urllib
import openpyxl
from datetime import datetime

#Realtime Data scrapper for Finazen Website. 

link = "http://www.finanzen.net/index/DAX-Realtime"




class StockScrapper(object):
    
    def __init__(self):
        self.driver = webdriver.PhantomJS()
        self.driver.set_window_size(1120, 550)
        self.pagesource = None
        self.headings = []
        self.driver.get(link)
        self.count = 0

    
    def getDAX(self):

        self.pagesource = BeautifulSoup(self.driver.page_source, "html.parser")  
        value = self.pagesource.find("div", attrs={"source" : "lightstreamer", "field":"bid", "class": "inline", "table": 4})

        return value.get_text()

    def cell_writer(self, DAX=None):
        #This function writes the cell headings as well as the realtime data
        if DAX:
            wb = openpyxl.load_workbook('test.xlsx')
            sheet = wb.active
            rowtoupdate = sheet.max_row + 1

            sheet.cell(row = rowtoupdate, column =1).value = datetime.now().time()
            sheet.cell(row = rowtoupdate, column = 2).value = DAX
            wb.save('test.xlsx')
        else:
            wb = openpyxl.Workbook()

            headings = ['Time', 'Value']
            sheet= wb.active
            for j in range(0, len(headings)):
                sheet.cell(row = 1,column=j+1).value = headings[j]

            wb.save("test.xlsx")

    def current(self):
        wb = openpyxl.load_workbook('test.xlsx')
        sheet = wb.active
        curr_row = sheet.max_row
        return sheet.cell(row =sheet.max_row, column = 2).value
    
if __name__ == '__main__':
    
    scraper = StockScrapper()
    
    #Writing Headings
    scraper.cell_writer()

    #Current Time
    now = datetime.now()
    
    #DAX runs between 12:30 and 21:00 IST therefore the script would run between that only.
    starttime = now.replace(hour=12, minute=30, second=0, microsecond=0) 
    endtime = now.replace(hour=22, minute=30, second=0, microsecond=0)
    
    #Looping to update the excel file.
    while ( True ):
        getcurrentvalue = scraper.getDAX()
        if getcurrentvalue != scraper.current():
            print getcurrentvalue
            scraper.cell_writer(getcurrentvalue)






