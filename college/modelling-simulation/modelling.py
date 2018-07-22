import openpyxl
import math

b = []


# for i in a:
# 	temp = i.get_text()
# 	temp = [x.strip() for x in temp.split('\n')]
# 	while '' in temp:
# 		temp.remove('')

# 	b.append(temp)


wb = openpyxl.load_workbook('specialMess.xlsx')
sheet = wb.active
maxrow = sheet.max_row

dinner = []
days = []

for index in range(maxrow):
	dinner.append(sheet.cell(row=index+1, column=3).value)
	days.append(sheet.cell(row=index+1, column=2).value)


del dinner[0]
del days[0]

for index, item in enumerate(dinner):
	dinner[index] = math.floor(item)


mon = []
tue = []
wed = []
thur = []
fri = []
sat = []
sun = []

for index, day in enumerate(days):
	if day == 'Mon':
		mon.append(dinner[index])
	elif day == 'Tue':
		tue.append(dinner[index])
	elif day == 'Wed':
		wed.append(dinner[index])
	elif day == 'Thu':
		thur.append(dinner[index])
	elif day == 'Fri':
		fri.append(dinner[index])
	elif day == 'Sat':
		sat.append(dinner[index])
	elif day == 'Sun':
		sun.append(dinner[index])

print "Mon->",mon
print "Tue->",tue
print "Wed->",wed
print "Thur->",thur
print "Fri->", fri
print "Sat->", sat
print "Sun->", sun


wb.close()
# for index, item in enumerate(headings):
# 	sheet.cell(row=rowtoupdate, column=index+1).value = item

# rowtoupdate = rowtoupdate + 1

# for x in b:


# 	if len(x)  == 3:
# 		sheet.cell(row = rowtoupdate, column =1).value = x[0]
# 		sheet.cell(row = rowtoupdate, column = 2).value = x[1]
# 		sheet.cell(row = rowtoupdate, column = 4).value = x[2]
# 	else:
# 		for index, item in enumerate(x):
# 			sheet.cell(row = rowtoupdate, column = index + 1).value = item

# 	rowtoupdate = rowtoupdate + 1



# wb.save('test.xlsx')

# with xlrd.open_workbook('test.xlsx') as wb:
#     sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
#     with open('test.csv', 'wb') as f:
#         c = csv.writer(f)
#         for r in range(sh.nrows):
#             c.writerow(sh.row_values(r))
