import grequests
import time
import datetime
import time
import openpyxl
import xlrd
import csv
import math
import decimal


netprofit_one = 0
netprofit_break = 0


def calculateProfit(higher, lower):
	factor = 1
	profit_one = higher - lower - lower*0.25/100 - 0.001
	profit_break = profit_one

	while profit_break < 0.01:
		profit_break = higher*factor - lower*factor - lower*0.25/100*factor - 0.001
		# print profit_break
		if profit_break < 0.01:
			factor+=1
	
	return profit_one, profit_break, factor

urls = [
	'https://bittrex.com/api/v1.1/public/getticker?market=BTC-VTC',
    'https://poloniex.com/public?command=returnTicker',
	# 'https://bleutrade.com/api/v2/public/getticker?market=VTC_BTC',
	# 'http://data.bter.com/api2/1/ticker/vtc_btc'
]

ten = time.time() + 60*1

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row
headings = ['Time', 'Bittrex', 'Poloniex', 'Profit (1 coin)', 'Coins needed for breakeven', 'Profit to breakeven']

for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item

rowtoupdate = rowtoupdate + 1

while time.time() < ten:
	rs = (grequests.get(u) for u in urls)
	a =  grequests.map(rs)
	BIT = a[0].json()['result']['Last']
	POL = a[1].json()['BTC_VTC']['last']
	
	# BLE = a[1].json()['result'][0]['Last']
	# BTER = a[2].json()['last']
	
	highest = max(float(POL),float(BIT))
	lowest = min(float(POL),float(BIT))
	profit_one, profit_break, factor = calculateProfit(float(highest),float(lowest))
	netprofit_one = netprofit_one + profit_one
	netprofit_break = netprofit_break + profit_break

	sheet.cell(row=rowtoupdate, column=1).value = time.ctime()
	sheet.cell(row=rowtoupdate, column=2).value = BIT
	sheet.cell(row=rowtoupdate, column=3).value = POL
	sheet.cell(row=rowtoupdate, column=4).value = profit_one
	sheet.cell(row=rowtoupdate, column=5).value = factor
	sheet.cell(row=rowtoupdate, column=6).value = profit_break

	rowtoupdate = rowtoupdate + 1
	

	print "--------------------"
	print time.ctime()
	print "Highest->", highest,
	print "Lowest->", lowest
	print "Profit (1 coin)->", profit_one
	print "Coins needed->", factor
	print "Profit Breakeven >=0.1->", profit_break
	
	
	time.sleep(30)

sheet.cell(row=rowtoupdate, column=4).value = netprofit_one
sheet.cell(row=rowtoupdate, column=6).value = netprofit_break
	
wb.save('status.xlsx')

with xlrd.open_workbook('status.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('status.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))
