import grequests
import time
import datetime
import time
import openpyxl
import xlrd
import csv
import math
import decimal


netprofit_one = 0
netprofit_break = 0


def calculateProfit(higher, lower, trade=0, network=0.001):
	factor = 1
	profit_one = higher - lower - lower*trade - network
	profit_break = profit_one
	
	print profit_break
	
	while profit_break < 0:
		
		profit_break = higher*factor - lower*factor - lower*trade*factor - network
		# print profit_break
		if profit_break < 0:
			factor+=1
	
	double = factor*2
	profit_double = higher*double - lower*double - lower*trade*double - network

	return factor, profit_break, profit_double

urls = [
	'https://api.hitbtc.com/api/1/public/ticker',
	'https://poloniex.com/public?command=returnTicker',
	'https://api.bitfinex.com/v1/pubticker/ltcbtc',
    'https://bitkonan.com/api/ltc_ticker',
	'https://bitkonan.com/api/ticker',
	# 'https://c-cex.com/t/ltc-btc.json'
]

ten = time.time() + 60*1

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row

headings = ['Time', 'HitBTC->Poloniex', 'Bitfinex->HitBTC', 'Bitkonan->HitBTC']
fees = [[0,0.001],[0.2/100,0],[0,0.001]]
c = ['Coins', 'Profit', '2xProfit']

counter = 1
for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=counter).value = item

	if index == 0:
		counter+=1
	else:
		for i, e in enumerate(c):
			sheet.cell(row=rowtoupdate+1, column=counter+i).value= e
		counter+=3

# for i in range(len(headings)):
# 	if not i == 0:
i = 2
while i < sheet.max_column:
	sheet.merge_cells(start_row=1,start_column=i, end_row=1, end_column=i+2)
	i+=3

rowtoupdate = rowtoupdate + 2

while time.time() < ten:
	rs = (grequests.get(u) for u in urls)
	a =  grequests.map(rs)
	HITBTC = float(a[0].json()['LTCBTC']['last'])
	POL = float(a[1].json()['BTC_LTC']['last'])
	BITFINEX = float(a[2].json()['last_price'])
	BITKONAN = float(a[3].json()['last'])/float(a[4].json()['last'])
	# CCEX = float(a[4].json()['lastprice'])


	sheet.cell(row=rowtoupdate, column=1).value = time.ctime()
	
	#Trades
	trades = [[HITBTC,POL], [BITFINEX,HITBTC], [BITKONAN, HITBTC]]

	j = 2
	fees_index = 0
	for i in trades:
		print i
		highest = max(i)
		print highest
		lowest = min(i)
		print lowest
		
		factor, profit_break, profit_double = calculateProfit(highest,lowest,fees[fees_index][0],fees[fees_index][1])
		
		sheet.cell(row=rowtoupdate, column=j).value = factor
		sheet.cell(row=rowtoupdate, column=j+1).value = profit_break
		sheet.cell(row=rowtoupdate, column=j+2).value = profit_double
		j+=3

	rowtoupdate = rowtoupdate + 1
	time.sleep(30)

# sheet.cell(row=rowtoupdate, column=4).value = netprofit_one
# sheet.cell(row=rowtoupdate, column=6).value = netprofit_break
	
wb.save('status.xlsx')

with xlrd.open_workbook('status.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('status.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))
