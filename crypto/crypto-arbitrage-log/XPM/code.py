import grequests
import time
import datetime
import time
import openpyxl
import xlrd
import csv

netprofit_one = 0
netprofit_break = 0


def calculateProfit(higher, lower):
	factor = 1
	profit_one = higher - lower - lower*0.25/100 - 0.001
	profit_break = profit_one

	while profit_break < 0:
		profit_break = higher*factor - lower*factor - lower*0.25/100*factor - 0.001
		if not profit_break >=0:
			factor+=1

	return profit_one, profit_break, factor	


urls = [
    'https://poloniex.com/public?command=returnTicker',
	'https://bleutrade.com/api/v2/public/getticker?market=XPM_BTC',
	'http://data.bter.com/api2/1/ticker/xpm_btc'
]



ten = time.time() + 60*10

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row
headings = ['Time', 'Poloniex', 'Bleutrade', 'Bter', 'Profit (1 coin)', 'Coins needed for breakeven', 'Profit to breakeven', '']

for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item

rowtoupdate = rowtoupdate + 1

while time.time() < ten:
	rs = (grequests.get(u) for u in urls)
	a =  grequests.map(rs)
	POL = a[0].json()['BTC_XPM']['last']
	BLE = a[1].json()['result'][0]['Last']
	BTER = a[2].json()['last']
	
	highest = max(float(POL),float(BLE))
	lowest = min(float(POL),float(BLE))
	profit_one, profit_break, factor = calculateProfit(float(highest),float(lowest))
	netprofit_one = netprofit_one + profit_one
	netprofit_break = netprofit_break + profit_break

	sheet.cell(row=rowtoupdate, column=1).value = time.ctime()
	sheet.cell(row=rowtoupdate, column=2).value = POL
	sheet.cell(row=rowtoupdate, column=3).value = BLE
	sheet.cell(row=rowtoupdate, column=4).value = BTER
	sheet.cell(row=rowtoupdate, column=5).value = profit_one
	sheet.cell(row=rowtoupdate, column=6).value = factor
	sheet.cell(row=rowtoupdate, column=7).value = profit_break

	rowtoupdate = rowtoupdate + 1
	print highest,lowest, profit_one, factor, profit_break
	print time.ctime()

	time.sleep(30)

sheet.cell(row=rowtoupdate, column=5).value = netprofit_one
sheet.cell(row=rowtoupdate, column=7).value = netprofit_break
	
wb.save('status.xlsx')

with xlrd.open_workbook('status.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('status.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))
