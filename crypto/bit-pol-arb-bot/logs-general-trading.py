import time
import openpyxl
import requests
from bittrex import bittrex
import poloniex
import sys
import math
from helper import keys, helperfunctions, currencies

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row

headings = ['Time', 'Currency', 'Higher', 'Lower', 'Coins', 'Margin', 'Status','BTC Required', 'BTC < 0.1', 'Dollar Required', 'Dollar Made']

for index, item in enumerate(headings):
    sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate=rowtoupdate+1

maximum = {}
minimum = {}

p = poloniex.Poloniex(keys.pol['key'],keys.pol['secret'])
b = bittrex.Bittrex(keys.bit['key'],keys.bit['secret'])

minBTC = []

def calculateProfit(higher, lower, hightrade=0, lowtrade=0, network=0):
        profit = higher - higher*hightrade - lower - lower*lowtrade - network
        factor = int(math.ceil(((1.009)*network)/(higher - 1.009*(lower + lower*lowtrade + higher*hightrade))))
        sp = higher*factor
        cp = factor*(lower + lower*lowtrade + higher*hightrade) + network 
        margin = ((sp - cp)/cp)*100
        if factor>=1:
            print factor, margin
            return True, factor, margin
        else:
            return False, factor, margin

def checkArb(currencyPair):
    print "Checking->", currencyPair

    global rowtoupdate, p, b
    sheet.cell(row=rowtoupdate, column=1).value = time.ctime()
    sheet.cell(row=rowtoupdate, column=2).value = currencyPair

    base = currencyPair.split('_')[0]
    currency = currencyPair.split('_')[1]
    
    try:
        p_c = float(p.returnTicker()[currencyPair]['last'])
        b_c = float(b.get_ticker(currencyPair.replace("_","-"))['result']['Last'])
        highest = max(p_c, b_c)
        lowest = min(p_c, b_c)

        if highest == p_c:
            sheet.cell(row=rowtoupdate, column=3).value = "POL " + str(p_c) 
            sheet.cell(row=rowtoupdate, column=4).value = "BIT " + str(b_c)
            
            trade, factor, margin = calculateProfit(highest,lowest,0.25/100, 0.25/100, helperfunctions.transferFees(currencyPair))

        else:
            sheet.cell(row=rowtoupdate, column=3).value = "BIT " + str(b_c) 
            sheet.cell(row=rowtoupdate, column=4).value = "POL " + str(p_c) 

            trade, factor, margin = calculateProfit(highest,lowest,0.25/100, 0.15/100, helperfunctions.transferFees(currencyPair))

        if trade:
            print "------------"
         
            sheet.cell(row=rowtoupdate, column=5).value = factor
            sheet.cell(row=rowtoupdate, column=7).value = "Trading"
            sheet.cell(row=rowtoupdate, column=6).value = margin

            if highest == p_c:
                print "Higher-> Pol", p_c*factor
                print "Lower-> Bit", b_c*factor
                print "Margin->", margin
                
                if base == "BTC":
                    if b_c*factor < 0.00081840:
                        print "-------------------------"
                        print "Checking->", currencyPair
                        print "TRADE!"
                        print "Higher-> Poloniex, Lower-> Bittrex "
                        print "ETH Needed->", b_c*factor
            
                dollar = float(b.get_ticker(checkBase(base).replace("_","-"))['result']['Last'])
                
                amount = dollar*b_c*factor
                made = dollar*b_c*margin/100*factor
              
               
                sheet.cell(row=rowtoupdate, column=8).value = b_c*factor
                if b_c*factor < 0.1:
                    sheet.cell(row=rowtoupdate, column=9).value = "YES"
                sheet.cell(row=rowtoupdate, column=10).value = amount
                sheet.cell(row=rowtoupdate, column=11).value = made
                
                minBTC.append((currencyPair, "BIT" , b_c*factor))
                # print "--------------------"
            
            else:
                
                print "Higher-> Bittrex", b_c*factor
                print "Lower-> Poloniex", p_c*factor
                print "Factor->", factor

                if factor < 30:
                    print "TRADE"
                
                if base == "ETH":
                    if p_c*factor < 0.03925556:
                        print "-------------------------"
                        print "Checking->", currencyPair
                        print "TRADE"
                        print "Higher-> Bittrex, Lower-> Poloniex "
                        print "ETH Needed->", p_c*factor
                elif base == "BTC":
                    if currency != "SJCX":
                        if p_c*factor < float(p.returnBalances()[base]):
                            print "-------------------------"
                            print "Checking->", currencyPair
                            print "TRADE"
                            print "Higher-> Bittrex, Lower-> Poloniex "
                            print "BTC Needed->", p_c*factor
                    
                    
                
                dollar = float(p.returnTicker()[checkBase(base)]['last'])
                
                amount = dollar*p_c*factor
                made = dollar*p_c*margin/100*factor

                
               
                sheet.cell(row=rowtoupdate, column=8).value = p_c*factor
                if p_c*factor < 0.1:
                    sheet.cell(row=rowtoupdate, column=9).value = "YES"
                sheet.cell(row=rowtoupdate, column=10).value = amount
                sheet.cell(row=rowtoupdate, column=11).value = made

                minBTC.append((currencyPair,"POL", p_c*factor))
                # print "--------------------"
            
        else:
            sheet.cell(row=rowtoupdate, column=5).value = "Factor Neg"
            sheet.cell(row=rowtoupdate, column=7).value = "Not Trading"
            sheet.cell(row=rowtoupdate, column=6).value = margin
        

        rowtoupdate+=1
        wb.save('logs/general-trading.xlsx')      
    except Exception as e:
        return e
    
   
if __name__ == "__main__":
    
    while 1:
        if len(sys.argv) == 1:
            print "GENERAL MODE"
            
            for currency in currencies.currencies_on_bit:
                checkArb(currency)

            # for i in range(11):
            #     if i + 1 != 8:
            #         sheet.cell(row=rowtoupdate, column=i+1).value = "MIN"
            # sheet.cell(row=rowtoupdate, column=8).value = sorted(minBTC, key=lambda x: x[2])[0][2]
            # print "MIN->", sorted(minBTC, key=lambda x: x[2])
            rowtoupdate+=1

        else:
            print sys.argv[1]
            checkArb(sys.argv[1])