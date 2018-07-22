import time
import openpyxl
import xlrd
import csv
import requests
import time
from bittrex import bittrex
import poloniex
import sys
import math
from helper import keys, helperfunctions, daddress, currency

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row

headings = ['Time', 'Currency', 'Higher', 'Lower', 'Coins', 'Margin', 'Status', 'Dollar Required', 'Dollar Made']

for index, item in enumerate(headings):
    sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate=rowtoupdate+1

p = poloniex.Poloniex(keys.pol['key'],keys.pol['secret'])
b = bittrex.Bittrex(keys.bit['key'],keys.bit['secret'])


def calculateProfit(higher, lower, hightrade=0, lowtrade=0, network=0):
        profit = higher - higher*hightrade - lower - lower*lowtrade - network
        factor = int(math.ceil(((1.009)*network)/(higher - 1.009*(lower + lower*lowtrade + higher*hightrade))))
        sp = higher*factor
        cp = factor*(lower + lower*lowtrade + higher*hightrade) + network 
        margin = ((sp - cp)/cp)*100
        if factor>=1 and factor<=1000:
            return True, factor, margin
        else:
            return False, factor, margin

def bittrexToPoloniex(currencyPair,base, currency,factor,b_c,p_c):
    print "Transfer from bittrex to poloniex"
    if b.get_balance(base)['result']['Balance'] > b_c*factor:
        print "Trade Happening"
        order = b.buy_limit(currencyPair.replace("_","-"),factor,b_c)
        if order["success"]:
            uuid = order['result']['uuid']
            while not BittrexBuySuccess(currencyPair,uuid):
                print "Order still open"

            print "Buy completed at bittrex"

            quantity = b.get_balance(base)['result']['Balance']

            if currency in depositAddressP:
                withdrawalUuid = b.withdraw(currencyPair,quantity,depositAddressP[currency])['result']['uuid']
                
                print "Withdrawn from bittrex"
                while 1:
                    txid = getTxid(withdrawalUuid)
                    if txid != False:
                        break;

                while not PoloniexDepositStatus(currency,txid):
                    print "Not recieved at poloniex"
                print "recieved at poloniex"
                sellOrder = p.sell(currencyPair,p_c,factor)
                print "Sell order placed at Poloniex"
                sellOrderNumber = sellOrder.orderNumber
                while not PoloniexOrderStatus(currencyPair,sellOrderNumber):
                    print "Order still open"
                print "Sell order successful at Poloniex"
            else:
                print "No currency address found for Poloniex"


        else:
            print "But not successful at bittrex"

    else:
        print "Factors do not meet trade"

def BittrexBuySuccess(currencyPair, uuid):
    openOrders = b.get_open_orders(currencyPair)["result"]
    for order in openOrders:
        if order["OrderUuid"] == uuid:
            return False
    return True

def PoloniexBuySuccess(currencyPair, orderNumber):
    openOrders = p.returnOpenOrders(currencyPair)
    for i in openOrders:
        if i['orderNumber'] == orderNumber:
            return False
    
    return True

def BittrexDepositStatus(currency):
    if float(b.get_balance(currency)['result']['Pending'])>0:
        return False
    else:
        return True

def PoloniexDepositStatus(txid):
    deposits = p.returnDepositsWithdrawals()["deposits"]
    for deposit in deposits:
        if deposit["txid"] == txid and deposit["status"] == "COMPLETE":
            return True
    return False

def BittrexOrderStatus(currencyPair, uuid):
    for i in b.get_open_orders(currencyPair)['result']:
        if i['OrderUuid'] == uuid:
            return False
    
    return True

def PoloniexOrderStatus(currencyPair, orderNumber):
    for order in p.returnOpenOrders(currencyPair):
        if order.orderNumber == orderNumber:
            return False
    return True

def getTxid(currencyPair,uuid):
    withdrawalHistory = b.get_withdrawal_history(currencyPair)['result']
    for i in withdrawalHistory:
        if i['PaymentUuid'] == uuid:
            return i['TxId']
    return False

def poloniexToBittrex(currencyPair, base, currency,factor,b_c,p_c):
    global rowtoupdate
    print "Transfer from poloniex to bittrex"
    if p.returnBalances()[base] < p_c*factor and float(p.returnTicker()['USDT_BTC']['baseVolume']) > factor:
        print "TRADE happening"
        order = p.buy(currencyPair,p_c,factor, orderType='fillOrKill')
        if order:
            orderNumber = order.orderNumber
            total = order['resultingTrades'][0].total
            while not PoloniexBuySuccess(currencyPair,orderNumber):
                print "Order still open"
            
            print "Buy Completed at Pol"
            

            quantity = p.returnBalances()[currency]

            if currency in depositAddressB:
                p.withdraw(currencyPair,quantity,depositAddressB[currency])
                print "Withdraw from Pol Happened"
                while not BittrexDepositStatus(currency):
                    print "Not recieved at Bit"
                
                print "Recieved at Bit"

                
                order = b.sell_limit(currencyPair.replace("_","-"),b.get_balance(currency)['result']['Balance'],b_c)
                uuid = order['result']['uuid']

                print "Sell order placed at Biitrex"

                while not BittrexOrderStatus(currencyPair.replace("_","-"), uuid):
                    print "Order still open"
            else:
                print "No deposit address found for Bittrex"
        else:
            "Buy not Successful on Poloniex"
        

    
    #    / time.sleep(20)

        

    else:
        print "Factors do not meet for Trade"
    

    # p.withdraw(currencyPair,quantity,'PS6GZr25Re1aukYNG4zbzR9siMsfzVyCB6')

    # expectedDeposit = quantity - checkBTCETH(currencyPair)
    # while b.get_balance(currency)['result']['Balance'] != quantity:
    #     print "Not Deposited at Bittrex"

    # b.sell_limit(currencyPair.replace("_","-"),b.get_balance(currency)['result']['Balance'],b_c)

def checkArb(currencyPair):
    
    global rowtoupdate, p, b
    sheet.cell(row=rowtoupdate, column=1).value = time.ctime()
    sheet.cell(row=rowtoupdate, column=2).value = currencyPair

    base = currencyPair.split('_')[0]
    currency = currencyPair.split('_')[1]
    
    try:
        pol = p.returnTicker()[currencyPair]
        bit = b.get_ticker(currencyPair.replace("_","-"))['result']['Last']
        
        p_c = float(pol['last'])
        b_c = float(bit)
    
    
        highest = max(p_c, b_c)
        lowest = min(p_c, b_c)

        if highest == p_c:
            sheet.cell(row=rowtoupdate, column=3).value = "POL " + str(p_c) 
            sheet.cell(row=rowtoupdate, column=4).value = "BIT " + str(b_c)
            print checkBTCETH(currencyPair)
            trade, factor, margin = calculateProfit(highest,lowest,0.25/100, 0.25/100, helperfunctions.transferFees(currencyPair))

        else:
            print checkBTCETH(currencyPair)
            sheet.cell(row=rowtoupdate, column=3).value = "BIT " + str(b_c) 
            sheet.cell(row=rowtoupdate, column=4).value = "POL " + str(p_c) 

            trade, factor, margin = calculateProfit(highest,lowest,0.25/100, 0.15/100, helperfunctions.transferFees(currencyPair))


        if trade:
            print "-------------------------"
            print "Checking->", currencyPair
            print "Trade with coins", factor

            sheet.cell(row=rowtoupdate, column=5).value = factor
            sheet.cell(row=rowtoupdate, column=7).value = "Trading"
            sheet.cell(row=rowtoupdate, column=6).value = margin

            if highest == p_c:
                print "Higher-> Poloniex", p_c*factor
                print "Lower-> Bittrex", b_c*factor
                print "Margin->", margin
            
                dollar = float(b.get_ticker(checkBase(base).replace("_","-"))['result']['Last'])
                
                amount = dollar*b_c
                made = dollar*b_c*margin/100

                print "Amount Needed->", amount
                print "Amount Made->", made

                sheet.cell(row=rowtoupdate, column=8).value = amount
                sheet.cell(row=rowtoupdate, column=9).value = made
                
                print "--------------------"
                bittrexToPoloniex(currencyPair,base, currency,factor,b_c,p_c)
            
            else:
                print "Higher-> Bittrex", b_c*factor
                print "Lower-> Poloniex", p_c*factor
                print "Margin->", margin
                
                dollar = float(p.returnTicker()[checkBase(base)]['last'])
                
                amount = dollar*p_c
                made = dollar*p_c*margin/100

                print "Amount Needed->", amount
                print "Amount Made->", made

                sheet.cell(row=rowtoupdate, column=8).value = amount
                sheet.cell(row=rowtoupdate, column=9).value = made

                print "--------------------"
                poloniexToBittrex(currencyPair,base, currency,factor,b_c,p_c)
            
        else:
            print "No Trading"
            sheet.cell(row=rowtoupdate, column=5).value = "FH"
            sheet.cell(row=rowtoupdate, column=7).value = "Not Trading"
            sheet.cell(row=rowtoupdate, column=6).value = margin

        rowtoupdate+=1 
        wb.save('logs/trade/trade.xlsx')
    except Exception as e:
        return e      


if len(sys.argv) == 2:
    c = sys.argv[1]
    if c not in currency.currencies_on_bit:
        print "Enter a valid currency from the following list \n", currency.currencies_on_bit
    else:
        checkArb(sys.argv[1])
else:
    print "specify the currency"