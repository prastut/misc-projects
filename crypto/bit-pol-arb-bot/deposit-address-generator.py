import time
import openpyxl
import xlrd
import csv
import requests
from bittrex import bittrex
import poloniex
import sys
import math
from helper import keys, currency, helperfunctions


wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row

headings = ['Exchange','Currency', 'Address']

for index, item in enumerate(headings):
    sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate=rowtoupdate+1


p = poloniex.Poloniex(keys.pol['key'],keys.pol['secret'])
b = bittrex.Bittrex(keys.bit['key'],keys.bit['secret'])

depositAddressB = {}

def generateDepositAddress(currency, exchange):
	global rowtoupdate
	for currencyPair in currency:
		base = currencyPair.split('_')[0]
		currency = currencyPair.split('_')[1]
			
		if not b.get_deposit_address(currency)['result'] == None:
			depositAddressB[b.get_deposit_address(currency)['result']['Currency']] = b.get_deposit_address(currency)['result']['Address']
			sheet.cell(row=rowtoupdate, column=1).value = exchange
			sheet.cell(row=rowtoupdate, column=2).value = currencyPair
			sheet.cell(row=rowtoupdate, column=3).value = depositAddressB[b.get_deposit_address(currency)['result']['Currency']]
			rowtoupdate+=1
        

		wb.save('logs/deposit-address-' + exchange + '.xlsx')   

		print depositAddressB


if __name__ == "__main__":
	generateDepositAddress(currency.currencies_on_bit, "Bittrex");
print depositAddressB