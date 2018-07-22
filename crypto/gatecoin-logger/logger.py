import requests as r
import openpyxl
import time
from datetime import datetime

filename = 'logs/log_' + datetime.now().strftime('%Y_%m_%d_%H_%M_%S') + '.xlsx'

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = 1


headings = ['Time', 'HKD-USD', 'BTC-HKD-USD', 'BTC-USD', 'Margin BTC', 'ETH-HKD-USD', 'ETH-USD', 'Margin ETH']


for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate +=1


while 1:
	
	try:

		sheet.cell(row=rowtoupdate, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
		

		gatecoin = r.get('https://api.gatecoin.com/Public/LiveTickers').json()

		for i in gatecoin['tickers']:
			if i['currencyPair'] == "BTCUSD":
				btc_usd = float(i['last'])
			elif i['currencyPair'] == "BTCHKD":
				btc_hkd = float(i['last'])
			elif i['currencyPair'] == "ETHHKD":
				eth_hkd = float(i['last'])
			elif i['currencyPair'] == "ETHUSD":
				eth_usd = float(i['last']) 

		hkd_usd =  float(r.get('http://api.fixer.io/latest?base=HKD').json()["rates"]["USD"])

		btc_hkd_usd = btc_hkd * hkd_usd
		eth_hkd_usd = eth_hkd * hkd_usd

		# BTC Margin
		low_btc  = min(btc_usd, btc_hkd_usd)
		high_btc = max(btc_usd, btc_hkd_usd)
		margin_btc = ((high_btc - low_btc)/low_btc)*100

		# ETH Margin
		low_eth  = min(eth_usd, eth_hkd_usd)
		high_eth = max(eth_usd, eth_hkd_usd)
		margin_eth = ((high_eth - low_eth)/low_eth)*100


		print ("BTC-HKD-USD {}, BTC-USD {}".format(btc_hkd_usd, btc_usd))

		sheet.cell(row=rowtoupdate, column=2).value = hkd_usd
		sheet.cell(row=rowtoupdate, column=3).value = btc_hkd_usd
		sheet.cell(row=rowtoupdate, column=4).value = btc_usd
		sheet.cell(row=rowtoupdate, column=5).value = margin_btc
		sheet.cell(row=rowtoupdate, column=6).value = eth_hkd_usd
		sheet.cell(row=rowtoupdate, column=7).value = eth_usd
		sheet.cell(row=rowtoupdate, column=8).value = margin_eth
		
		
		rowtoupdate+=1

		wb.save(filename)
		
		
	except Exception as e:
		print(e)
		pass
	
	time.sleep(60*5)
