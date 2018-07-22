import json

from datetime import datetime
import openpyxl
import dateparser


wb = openpyxl.load_workbook('Kraken_USD_BTC.xlsx')
data_sheet = wb.active
row_max = data_sheet.max_row

write_log = openpyxl.Workbook()
sheet = write_log.active

dates = []

for rows in range(row_max):

	date = dateparser.parse(str(data_sheet.cell(row=rows+2, column=1).value))
	
	if date:
		dates.append(date.strftime('%Y-%m-%d'))


set_dates = sorted(set(dates))


for date in set_dates:
	write_log.create_sheet(date)

	sheet = write_log[date]
	rowtoupdate = 1
	headings = ['Time', 'AVG']

	for index, item in enumerate(headings):
		sheet.cell(row=rowtoupdate, column=index+1).value = item
	
	rowtoupdate+=1

	for row in range(row_max):

		loop_date = dateparser.parse(str(data_sheet.cell(row=row+1, column=1).value))
	
		if loop_date:
			if loop_date.strftime('%Y-%m-%d') == date:
				sheet.cell(row=rowtoupdate, column=1).value = loop_date
				sheet.cell(row=rowtoupdate, column=2).value = str(data_sheet.cell(row=row+1, column=8).value)
				rowtoupdate+=1
				

		# try:
			
		# 	if dateparser.parse(str(data_sheet.cell(row=rows+2, column=1).value)).strftime('%Y-%m-%d')== date:
		# 		print (dates[row], date)
		# 		
			
		# except:
		# 	pass
		
	
	print ("New Loop")

write_log.save('kraken.xlsx')