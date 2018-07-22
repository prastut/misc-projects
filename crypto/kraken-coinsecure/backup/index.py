import json
from coin_secure import data as cs
import datetime
import openpyxl


wb = openpyxl.Workbook()
sheet = wb.active




dates = []

for row in cs[0]['monthly'][0]:	
	dates.append(datetime.datetime.fromtimestamp(int(row['time'])/1000).strftime('%Y-%m-%d'))

dates_bucket = sorted(set(dates))


for date in dates_bucket:
	wb.create_sheet(date)

	date_sheet = wb[date]
	rowtoupdate = 1
	headings = ['Time', 'AVG']

	for index, item in enumerate(headings):
		date_sheet.cell(row=rowtoupdate, column=index+1).value = item
	
	rowtoupdate+=1

	for row in cs[0]['monthly'][0]:
		
		if datetime.datetime.fromtimestamp(int(row['time'])/1000).strftime('%Y-%m-%d') == date:
			date_sheet.cell(row=rowtoupdate, column=1).value = datetime.datetime.fromtimestamp(int(row['time'])/1000).strftime('%Y-%m-%d  %H:%M:%S')
			date_sheet.cell(row=rowtoupdate, column=2).value = row['avg']
			rowtoupdate+=1
		



wb.save('coinsecure.xlsx')