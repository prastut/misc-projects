from websocket import create_connection
import time, json
import openpyxl
from datetime import datetime
import requests


wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = 1
headings = ['Time', 'Kraken USD']


for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate +=1


print("Receiving...")

while 1:
	
	try:
	

		sheet.cell(row=rowtoupdate, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

		kraken = requests.get('https://api.kraken.com/0/public/Ticker', params={'pair': 'XXBTZUSD'})
		json_dump = kraken.json()
		sheet.cell(row=rowtoupdate, column=2).value = json_dump['result']['XXBTZUSD']['c'][0]
		
		print ("Time: {}, Price {}".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'),json_dump['result']['XXBTZUSD']['c'][0])) 
		wb.save('../logs/daily/kraken.xlsx')
		rowtoupdate+=1
	except Exception as e:
		print (e)

	print ("Delaying")
	time.sleep(60*60)