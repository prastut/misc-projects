from websocket import create_connection
import time, json
import openpyxl
from datetime import datetime

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = 1
headings = ['Time', 'Coinsecure INR']


for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate +=1


print("Receiving...")

while 1:
	
	try:
		ws = create_connection("wss://coinsecure.in/websocket")
		result =  ws.recv()
		parsed = json.loads(str(result))
		lastPrice = parsed[0]["ticker"][3][0]["lasttrade"][0]
		# print(lastPrice)
		sheet.cell(row=rowtoupdate, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

	
		if 'ask' in lastPrice:	
			sheet.cell(row=rowtoupdate, column=2).value = lastPrice['ask'][0][0]['rate']
			print ("Time: {}, Price {}".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'),lastPrice['ask'][0][0]['rate'])) 
		elif 'bid' in lastPrice:
			sheet.cell(row=rowtoupdate, column=2).value = lastPrice['bid'][0][0]['rate']
			print ("Time: {}, Price {}".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'),lastPrice['bid'][0][0]['rate'])) 

		rowtoupdate+=1	
		wb.save('../logs/daily/coinsecure.xlsx')
		ws.close()
		
	except Exception as e:
		print e

	print ("Delaying")
	time.sleep(60*60)