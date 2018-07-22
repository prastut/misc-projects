import requests as r
import openpyxl
from bittrex import bittrex
import keys
import time
from datetime import datetime
#ANS / Ant Shares


# Yunbi 

filename = 'logs/daily/log_' + datetime.now().strftime('%Y_%m_%d_%H_%M_%S') + '.xlsx'

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = 1


headings = ['Time', 'Yunbi USD-CNY-ANS', 'Bittrex USDT-BTC-ANS']


for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate +=1


b = bittrex.Bittrex(keys.bit['key'],keys.bit['secret'])

while 1:
	
	try:

		sheet.cell(row=rowtoupdate, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
		
		# Yunbi
		yunbi_all = r.get("https://yunbi.com//api/v2/tickers.json")
		yunbi_yen = float(yunbi_all.json()["anscny"]['ticker']['last'])
		
		yen_to_usd_API = r.get('http://api.fixer.io/latest?base=CNY')
		yen_to_usd = float(yen_to_usd_API.json()["rates"]["USD"])
		
		yunbi_usd = yunbi_yen * yen_to_usd

		# Bittrex

		bit_BTC_ANS = float(b.get_ticker('BTC-ANS')['result']['Last'])
		bit_USD_BTC = float(b.get_ticker('USDT-BTC')['result']['Last'])

		bit_usd = bit_BTC_ANS * bit_USD_BTC


		print(yunbi_usd, bit_usd)
		
		sheet.cell(row=rowtoupdate, column=2).value = yunbi_usd
		sheet.cell(row=rowtoupdate, column=3).value = bit_usd

		rowtoupdate+=1

		wb.save(filename)
		
		
	except Exception as e:
		print(e)
		pass
	
	time.sleep(60*60)
