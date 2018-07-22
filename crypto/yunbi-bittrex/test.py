from websocket import create_connection
import time, json
import openpyxl
from datetime import datetime
import requests

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = 1
headings = ['Time', 'Kraken USD', 'Coinsecure INR']


for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item
rowtoupdate +=1


print("Receiving...")

while 1:
	
	try:
		ws = create_connection("wss://coinsecure.in/websocket")
		result =  ws.recv()
		parsed = json.loads(str(result))
		lastPrice = parsed[0]["ticker"][3][0]["lasttrade"][0]
		# print(lastPrice)
		sheet.cell(row=rowtoupdate, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
		
		kraken = requests.get('https://api.kraken.com/0/public/Ticker', params={'pair': 'XXBTZUSD'})
		json_dump = kraken.json()
		
		sheet.cell(row=rowtoupdate, column=2).value = json_dump['result']['XXBTZUSD']['c'][0]
		if 'ask' in lastPrice:	
			sheet.cell(row=rowtoupdate, column=3).value = lastPrice['ask'][0][0]['rate']
			print (lastPrice['ask'][0][0]['rate'])
		elif 'bid' in lastPrice:
			sheet.cell(row=rowtoupdate, column=3).value = lastPrice['bid'][0][0]['rate']
			print (lastPrice['bid'][0][0]['rate'])

		
		rowtoupdate+=1

		wb.save('both_2.xlsx')
		ws.close()
		
	except Exception as e:
		print(e)
		pass
	
	time.sleep(60*60)
