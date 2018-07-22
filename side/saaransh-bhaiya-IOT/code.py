import requests
from bs4 import BeautifulSoup
import openpyxl
import xlrd
import csv

b = []

page = requests.get("http://iotclouddata.com/project/598/iot16view.php")
soup = BeautifulSoup(page.content, 'html.parser')


a = soup.find_all('tr', class_='odd')

for i in a:
	temp = i.get_text()
	temp = [x.strip() for x in temp.split('\n')]
	while '' in temp:
		temp.remove('')

	b.append(temp)


wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row


headings = ['sno', 'traffic', 'date', 'time']

for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item

rowtoupdate = rowtoupdate + 1

for x in b:


	if len(x)  == 3:
		sheet.cell(row = rowtoupdate, column =1).value = x[0]
		sheet.cell(row = rowtoupdate, column = 2).value = x[1]
		sheet.cell(row = rowtoupdate, column = 4).value = x[2]
	else:
		for index, item in enumerate(x):
			sheet.cell(row = rowtoupdate, column = index + 1).value = item

	rowtoupdate = rowtoupdate + 1



wb.save('test.xlsx')

with xlrd.open_workbook('test.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('test.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))
