import requests
from bs4 import BeautifulSoup
import openpyxl
import xlrd
import csv
import re
from collections import defaultdict

page = requests.get("http://www.espncricinfo.com/indian-premier-league-2017/engine/series/1078425.html")
soup = BeautifulSoup(page.content, 'html.parser')



matchscrap = soup.find_all('span', class_='potMatchLink')
winscrap = soup.find_all('span', class_='show-for-small')

# Matchscrap

team1 = []
team2 = []
location = []

for i in matchscrap:
	temp = i.get_text()
	temp = [x.strip() for x in temp.split('\n')]
	while '' in temp:
		temp.remove('')
	
	temp = [x.strip() for x in re.split(r'\b(at|v)\b',temp[0])]
	team1.append(temp[0])
	team2.append(temp[2])
	location.append(temp[4])


# Winscrap

winner = []

for i in winscrap:
	temp = i.get_text()
	temp = [x.strip() for x in temp.split('\n')]
	while '' in temp:
		temp.remove('')
	
	temp = [x.strip() for x in re.split(r'\b(won)\b',temp[0])]
	winner.append(temp[0])

    	

# Populating score list:
teams = set(team1)
matchPos = {}
pointDict = {}

for team in teams:
	t1 = [i for i,x in enumerate(team1) if x==team]
	t2 = [i for i,x in enumerate(team2) if x==team]
	t1.extend(t2)
	t1.sort()
	
	#Match Position Dict
	matchPos[team]= t1

	pointcount = 0
	pointDict[team] = [pointcount]
	for pos in matchPos[team]:
		if winner[pos] == team:
			pointcount+=2
			pointDict[team].append(pointcount)
		else:
			pointDict[team].append(pointcount)




wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row


for index, key in enumerate(pointDict):
	sheet.cell(row=rowtoupdate, column=index+2).value = key
	for point in pointDict[key]:
		sheet.cell(row=rowtoupdate+1, column=index+2).value = point
		
		rowtoupdate+=1

	rowtoupdate = 1		

sheet.cell(row=1, column=1).value = 'match'

for index in range(sheet.max_row):
	sheet.cell(row=index+2, column=1).value = index + 1
	
	

wb.save('scorelist.xlsx')

# Status List:

wb = openpyxl.Workbook()
sheet = wb.active
rowtoupdate = sheet.max_row


headings = ['MatchID', 'Team1', 'Team2', 'Location', 'Winner']

for index, item in enumerate(headings):
	sheet.cell(row=rowtoupdate, column=index+1).value = item

rowtoupdate = rowtoupdate + 1

for index in range(len(winner)):
	sheet.cell(row = rowtoupdate, column = 1).value = index + 1
	sheet.cell(row = rowtoupdate, column = 2).value = team1[index]
	sheet.cell(row = rowtoupdate, column = 3).value = team2[index]
	sheet.cell(row = rowtoupdate, column = 4).value = location[index]
	sheet.cell(row = rowtoupdate, column = 5).value = winner[index]

	rowtoupdate = rowtoupdate + 1



wb.save('status.xlsx')

with xlrd.open_workbook('status.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('status.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))

with xlrd.open_workbook('scorelist.xlsx') as wb:
    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    with open('scorelist.csv', 'wb') as f:
        c = csv.writer(f)
        for r in range(sh.nrows):
            c.writerow(sh.row_values(r))
